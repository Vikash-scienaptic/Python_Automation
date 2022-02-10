from openpyxl import Workbook, load_workbook
import datetime
import csv

def func(file_name):
    wb = load_workbook(file_name)
    ws = wb['Q_Excel_PastDue']

    ws.delete_cols(21,24)
    ws.delete_cols(16,2)
    ws.delete_cols(14)
    ws.delete_cols(10,3)
    ws.delete_cols(9)
    ws.delete_cols(7)
    ws.delete_cols(5)
    ws.delete_cols(2)


    ws['A1'].value = 'Stock number'
    ws['B1'].value = 'name'
    ws['C1'].value = 'Home Phone'
    ws['D1'].value = 'Cell Phone'
    ws['E1'].value = 'Past Due amount'
    ws['F1'].value = 'days late'
    ws['G1'].value = 'portfolio'
    ws['H1'].value = 'City'
    ws['I1'].value = 'State'
    ws['J1'].value = 'Zip Code'

    for cell in ws['D']:
        if cell.value in ['',None]:
            a = cell.row
            cell.value = ws['C'][a-1].value

    for cell in ws['D']:
        if cell.value in ['',None]:
            ws.delete_rows(cell.row)

    ws.delete_cols(3)

    for cell in ws['B'][1:]:
        try:
            cell.value = cell.value.split(' ')[1] + " " + cell.value.split(' ')[0]
        except Exception as e:
            print(e)

    current_date = datetime.datetime.now().strftime("%Y%m%d")
    output_file = 'dialer_input_' + current_date 



    
    with open(output_file+'.csv','w',newline='') as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])

file_name = input(' enter the excel file name')



func(file_name)